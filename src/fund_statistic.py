# -*- coding:UTF-8 -*-

'''
Desc: 从基金的持仓中统计股票出现频率
File: /index.py
Project: src
File Created: Monday, 22nd March 2021 12:08:36 am
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2020 Camel Lu
'''
import time
import re
from pprint import pprint
import pandas as pd
from fund_info.statistic import FundStatistic
from utils.index import get_last_quarter_str, find_from_list_of_dict
from openpyxl import load_workbook
import os

def get_fund_code_pool():
    # fund_code_pool = ['000001', '160133', '360014', '420002',
    #                   '420102', '000409', '000418', '000746',
    #                   '000751', '000884', '000991', '001043',
    #                   '001054', '001104', '001410', '001473',
    #                   '519714', '000003', '000011', '000029']
    morning_star_rating_5_condition = {
        'value': 4,
        'operator': '>='
    }
    morning_star_rating_3_condition = {
        'value': 5,
        'operator': '='
    }
    last_year_time = time.localtime(time.time() - 365 * 24 * 3600)
    last_year_date = time.strftime('%Y-%m-%d', last_year_time)
    condition_dict = {
        'morning_star_rating_5': morning_star_rating_5_condition,
        'morning_star_rating_3': morning_star_rating_3_condition,
        # 'manager_start_date': '2020-05-25'
    }
    fund_code_pool = each_statistic.select_fund_pool(
        **condition_dict,
    )
    return fund_code_pool

def stocks_compare(stock_list, *, quarter_index=None, fund_code_pool=None, is_A_stock=None):
    each_statistic = FundStatistic()
    if quarter_index == None:
        quarter_index = get_last_quarter_str(2)

    filter_list = []
    for stock in stock_list:
        stock_code = stock[0].split('-', 1)[0]
        if not stock_code:
            continue
        stock_name = stock[0].split('-', 1)[1]
        stock_holder_detail = stock[1]
        holder_count = stock_holder_detail.get('count')
        holder_asset = stock_holder_detail.get('holder_asset')

        last_quarter_holder_detail_dict = each_statistic.select_special_stock_special_quarter_info(
            stock_code,
            quarter_index,
            fund_code_pool
        )
        
        last_holder_count = last_quarter_holder_detail_dict['count']
        last_holder_asset = last_quarter_holder_detail_dict['holder_asset']

        diff_holder_count = holder_count - last_holder_count
        diff_holder_asset = holder_asset - last_holder_asset

        diff_holder_count_percent = '{:.2%}'.format(
            diff_holder_count / last_holder_count) if last_holder_count != 0 else "+∞"
        
        diff_holder_asset_percent = '{:.2%}'.format(
            diff_holder_asset / last_holder_asset) if last_holder_asset != 0 else "+∞"
        # flag = '📈' if diff_holder_count > 0 else '📉'
        # if diff_holder_count == 0:
        #     flag = '⏸'
        flag_count = 'up' if diff_holder_count > 0 else 'down'
        if diff_holder_count == 0:
            flag = '='
        flag_asset = 'up' if diff_holder_asset > 0 else 'down'
        if diff_holder_asset == 0:
            flag = '='
            
        item_tuple = [stock_code, stock_name, holder_count, last_holder_count,
                      diff_holder_count, diff_holder_count_percent, flag_count, holder_asset, last_holder_asset, diff_holder_asset, diff_holder_asset_percent, flag_asset]
        if is_A_stock:
            industry_name_third = stock_holder_detail.get('industry_name_third')
            industry_name_second = stock_holder_detail.get('industry_name_second')
            industry_name_first = stock_holder_detail.get('industry_name_first')
            item_tuple = [*item_tuple, industry_name_third, industry_name_second,industry_name_first ]

        # if diff_percent == "+∞" or not float(diff_percent.rstrip('%')) < -20:
        filter_list.append(item_tuple)
        # print(item_tuple)
    return filter_list

# T100权重股排名
def t100_stocks_rank(quarter_index=None, *, each_statistic):
    if quarter_index == None:
        quarter_index = get_last_quarter_str()
    last_quarter_index = get_last_quarter_str(2)
    output_file = './outcome/数据整理/strategy/top100_rank.xlsx'
    sheet_name = quarter_index + '基金重仓股T100'
    columns=['代码',
        '名称', quarter_index + '持有数量（只）', last_quarter_index +'持有数量（只）', '持有数量环比', '持有数量环比百分比', '持有数量升或降',  quarter_index + '持有市值（亿元）', last_quarter_index + '持有市值（亿元）', '持有市值环比', '持有市值环比百分比', '持有市值升或降']

    stock_top_list = each_statistic.all_stock_fund_count(
        quarter_index=quarter_index,
        filter_count=80)
    stock_top_list = stock_top_list[:100]  # 获取top100权重股
    #pprint(stock_top_list)
    filter_list = stocks_compare(stock_top_list)
    df_filter_list = pd.DataFrame(filter_list, columns=columns)
    df_filter_list.to_excel(output_file, sheet_name=sheet_name)

# 所有股票排名
def all_stocks_rank(each_statistic):
    quarter_index = get_last_quarter_str(2)
    print("quarter_index", quarter_index)
    last_quarter_index = get_last_quarter_str(3)
    sheet_name = last_quarter_index + '基金重仓股T100'
    columns=['代码',
        '名称', quarter_index + '持有数量（只）', last_quarter_index +'持有数量（只）', '持有数量环比', '持有数量环比百分比', '持有数量升或降',  quarter_index + '持有市值（亿元）', last_quarter_index + '持有市值（亿元）', '持有市值环比', '持有市值环比百分比', '持有市值升或降']
    output_file = './outcome/数据整理/strategy/all_stock_rank/'+ quarter_index +'.xlsx'


    stock_top_list = each_statistic.all_stock_fund_count(
        quarter_index=quarter_index,
        filter_count=0)
    #print("stock_top_list", stock_top_list)
    all_a_stocks_industry_info_list = each_statistic.query_all_stock_industry_info()
    a_stock_list = []
    hk_stock_list = []
    other_stock_list = []
    for stock_name_code in stock_top_list:
        stock_code = stock_name_code[0].split('-', 1)[0]

        #path = 'other'
        if bool(re.search("^\d{5}$", stock_code)):
            #path = '港股'
            hk_stock_list.append(stock_name_code)
        elif bool(re.search("^\d{6}$", stock_code)):
            #'A股/深证主板'、'A股/创业板'、'A股/上证主板'、'A股/科创板'
            a_condition = bool(re.search("^(00(0|1|2|3)\d{3})|(30(0|1)\d{3})|(60(0|1|2|3|5)\d{3})|68(8|9)\d{3}$", stock_code))
            target_item = find_from_list_of_dict(all_a_stocks_industry_info_list, 'stock_code', stock_code)
            if a_condition and target_item:
                print('stock_code',stock_code)
                stock_name_code[1]['industry_name_first'] = target_item.get('industry_name_first')
                stock_name_code[1]['industry_name_second'] = target_item.get('industry_name_second')
                stock_name_code[1]['industry_name_third'] = target_item.get('industry_name_third')
                a_stock_list.append(stock_name_code)
            else:
                other_stock_list.append(stock_name_code)
        else:
            other_stock_list.append(stock_name_code)

    a_stock_compare_list = stocks_compare(a_stock_list, quarter_index=last_quarter_index, is_A_stock=True)
    hk_stock_compare_list = stocks_compare(hk_stock_list,quarter_index=last_quarter_index,)
    other_stock_compare_list = stocks_compare(other_stock_list,quarter_index=last_quarter_index,)
    a_columns = [*columns, '三级行业', '二级行业', '一级行业']
    df_a_list = pd.DataFrame(a_stock_compare_list, columns=a_columns)
    df_hk_list = pd.DataFrame(hk_stock_compare_list, columns=columns)
    df_other_list = pd.DataFrame(other_stock_compare_list, columns=columns)

    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    df_a_list.to_excel(writer, sheet_name='A股')

    df_hk_list.to_excel(writer, sheet_name='港股')

    df_other_list.to_excel(writer, sheet_name='其他')

    writer.save()

def all_stock_holder_detail(quarter_index, each_statistic, threshold=0):
    stock_list = each_statistic.all_stock_fund_count_and_details(
        quarter_index=quarter_index,
        filter_count=threshold)
    for i in range(0, len(stock_list)):
        stock = stock_list[i]
        stock_name_code = stock[0]
        stock_code = stock_name_code.split('-', 1)[0]
        path = '其他'
        if bool(re.search("^\d{5}$", stock_code)):
            path = '港股'
        elif bool(re.search("^\d{6}$", stock_code)):
            if bool(re.search("^00(0|1|2|3)\d{3}$", stock_code)):
                path = 'A股/深证主板'
            elif bool(re.search("^300\d{3}$", stock_code)):
                path = 'A股/创业板'
            elif bool(re.search("^60(0|1|2|3|5)\d{3}$", stock_code)):
                path = 'A股/上证主板'
            elif bool(re.search("^68(8|9)\d{3}$", stock_code)):
                path = 'A股/科创板'
            else:
                print('stock_name_code', stock_name_code)
        hold_fund_count = stock[1]['count']
        hold_fund_list = sorted(stock[1]['fund_list'], key=lambda x: x['持有市值(亿元)'], reverse=True) 
        df_list = pd.DataFrame(hold_fund_list)
        #if stock_code == 'NTES':
        #    print('stock_code', df_list)
        stock_name_code = stock_name_code.replace('-*', '-').replace('/', '-')
        path = './outcome/数据整理/stocks/' + path + '/' + stock_name_code + '.xlsx'
        path = path.replace('\/', '-')
        print("path", path)
        #print('df_list--',stock_name_code, df_list)
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(path)
            # 表名重复，删掉，重写
            if quarter_index in book.sheetnames:
                del book[quarter_index]
            if len(book.sheetnames) == 0:
                df_list.to_excel(
                path, sheet_name=quarter_index)
                continue
            else:
                writer.book = book
                df_list.to_excel(
                    writer, sheet_name=quarter_index)
            writer.save()
            writer.close()
        else:
            df_list.to_excel(
                path, sheet_name=quarter_index)
# 获取某些基金的十大持仓股票信息
def get_special_fund_code_holder_stock_detail(quarter_index, each_statistic):
    #基金组合信息
    fund_portfolio ={
        '001811': {
            'name': '中欧明睿新常态混合A',
            'position' : 0.2
        },
        '001705': {
            'name': '泓德战略转型股票',
            'position' : 0.2
        },
        '163415': {
            'name': '兴全商业模式优选混合',
            'position' : 0.2
        },
        '001043': {
            'name': '工银美丽城镇主题股票A',
            'position' : 0.2
        },
        '000547': {
            'name': '建信健康民生混合',
            'position' : 0.2
        },
         '450001': {
            'name': '国富中国收益混合',
            'position' : 0.2
        },
    }
    fund_code_pool = list(fund_portfolio.keys())
    holder_stock_industry_list = each_statistic.summary_special_funds_stock_detail(fund_code_pool, quarter_index)
    path = './outcome/数据整理/funds/' + '/' + '高分权益基金组合十大持仓明细' + '.xlsx'
    columns=['基金代码','基金名称', '基金类型', '基金经理', '基金总资产（亿元）', '基金股票总仓位', '十大股票仓位', '股票代码', '股票名称', '所占仓位', '所处仓位排名',  '三级行业', '二级行业', '一级行业']
    df_a_list = pd.DataFrame(holder_stock_industry_list, columns=columns)

    writer = pd.ExcelWriter(path, engine='xlsxwriter')
    df_a_list.to_excel(writer, sheet_name='十大持仓明细--' + quarter_index)
    writer.save()
if __name__ == '__main__':
    each_statistic = FundStatistic()
    quarter_index = "2021-Q1"

    # 所有股票的基金持仓细节
    #all_stock_holder_detail(quarter_index, each_statistic)

    # 获取所有股票排名,按股票市场分类输出
    all_stocks_rank(each_statistic)

    # 获取Top100股票排名
    #t100_stocks_rank(each_statistic=each_statistic)
    
    # 获取某些基金的十大持仓股票信息
    #get_special_fund_code_holder_stock_detail(quarter_index, each_statistic)
